"""
╔══════════════════════════════════════════════════════════════╗
║       ORKESTIA — Sistema de Gestión de Tareas               ║
║       Orquesta tu operación con Inteligencia                ║
║       Python + CustomTkinter  v4                            ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, sys, time, pickle, json, datetime, calendar, uuid, platform
import subprocess, ctypes, shutil, math
from pathlib import Path

try: import customtkinter as ctk
except ImportError:
    os.system(f"{sys.executable} -m pip install customtkinter"); import customtkinter as ctk
try: from tabulate import tabulate
except ImportError:
    os.system(f"{sys.executable} -m pip install tabulate"); from tabulate import tabulate
try: from tkcalendar import Calendar
except ImportError:
    os.system(f"{sys.executable} -m pip install tkcalendar"); from tkcalendar import Calendar
try:
    from openpyxl import Workbook as XlWb
    from openpyxl.styles import Font as XlF, PatternFill as XlP, Alignment as XlA
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook as XlWb
    from openpyxl.styles import Font as XlF, PatternFill as XlP, Alignment as XlA
try: from PIL import Image, ImageTk, ImageDraw
except ImportError:
    os.system(f"{sys.executable} -m pip install Pillow"); from PIL import Image, ImageTk, ImageDraw

from tkinter import messagebox, filedialog, Canvas
import tkinter as tk

# ─── Dirs ───
BASE = Path(__file__).parent
DATA = BASE/"data"; RPTS = BASE/"reportes"; BKUP = BASE/"backups"
AVDIR = DATA/"avatars"; FDIR = BASE/"fonts"; LOGO = BASE/"logo.png"
for d in [DATA,RPTS,BKUP,AVDIR,FDIR]: os.makedirs(d, exist_ok=True)
TFILE = DATA/"tareas.txt"; BFILE = DATA/"tareas.pkl"

# ─── Font ───
POP = False
def _reg(r):
    s=platform.system()
    try:
        if s=="Windows": return ctypes.windll.gdi32.AddFontResourceExW(str(r),0x10,0)>0
        elif s=="Darwin":
            d=Path.home()/"Library"/"Fonts"/Path(r).name
            if not d.exists(): shutil.copy2(r,d)
            return True
        else:
            fd=Path.home()/".local"/"share"/"fonts"; fd.mkdir(parents=True,exist_ok=True)
            d=fd/Path(r).name
            if not d.exists(): shutil.copy2(r,d)
            subprocess.run(["fc-cache","-f"],capture_output=True); return True
    except: pass
    return False
def _load_pop():
    global POP
    fs={"Poppins-Regular.ttf":"https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-Regular.ttf",
        "Poppins-Bold.ttf":"https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-Bold.ttf",
        "Poppins-SemiBold.ttf":"https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-SemiBold.ttf",
        "Poppins-Medium.ttf":"https://github.com/google/fonts/raw/main/ofl/poppins/Poppins-Medium.ttf"}
    if not all((FDIR/n).exists() for n in fs):
        try:
            import urllib.request
            for n,u in fs.items():
                r=FDIR/n
                if not r.exists(): urllib.request.urlretrieve(u,r)
        except: return
    ok=sum(_reg(str(FDIR/n)) for n in fs if (FDIR/n).exists())
    if ok>0: POP=True
_load_pop()
FF="Poppins" if POP else ("Helvetica Neue" if platform.system()=="Darwin" else "Segoe UI")

# ─── Colors ───
C={"bg":"#FFFFFF","bg2":"#F5F6F8","sb":"#FFFFFF","cd":"#FFFFFF","cdh":"#F0F1F3",
   "inp":"#F5F6F8","brd":"#E6E9EF","brdl":"#EEEFF2",
   "pri":"#6C60FF","prih":"#5B4FE8","pril":"#EDEDFF","pris":"#F0EEFF",
   "tx":"#323338","tx2":"#676879","tx3":"#9699A3","wh":"#FFFFFF",
   "ok":"#00CA72","okl":"#E6FAF0","wa":"#FDAB3D","wal":"#FFF5E6",
   "er":"#E2445C","erl":"#FDECEF","inf":"#579BFC","infl":"#E8F0FE",
   "pc":"#E2445C","pa":"#FDAB3D","pm":"#579BFC","pb":"#00CA72",
   "sp":"#FDAB3D","sg":"#579BFC","sc":"#00CA72",
   "c1":"#6C60FF","c2":"#00CA72","c3":"#FDAB3D","c4":"#E2445C","c5":"#579BFC","c6":"#FF6B9D"}

# ═══════════════ TOOLTIP ═══════════════
class Tooltip:
    def __init__(s,w,t,delay=400):
        s.w=w; s.t=t; s.delay=delay; s.tip=None; s._id=None
        w.bind("<Enter>",s._in); w.bind("<Leave>",s._out)
    def _in(s,e=None): s._id=s.w.after(s.delay,s._show)
    def _out(s,e=None):
        if s._id: s.w.after_cancel(s._id); s._id=None
        s._hide()
    def _show(s):
        if s.tip: return
        x=s.w.winfo_rootx()+s.w.winfo_width()//2; y=s.w.winfo_rooty()+s.w.winfo_height()+4
        s.tip=tw=tk.Toplevel(s.w); tw.wm_overrideredirect(True); tw.wm_geometry(f"+{x}+{y}")
        tk.Label(tw,text=s.t,bg="#323338",fg="white",relief="flat",bd=0,padx=8,pady=4,font=(FF,10)).pack()
    def _hide(s):
        if s.tip: s.tip.destroy(); s.tip=None

# ═══════════════ CALENDAR POPUP ═══════════════
class CalendarPopup:
    """Abre un Toplevel con un Calendar widget — funcional en todos los OS."""
    def __init__(self, parent, entry_var, initial_date=None):
        self.entry_var = entry_var
        self.result = None
        top = tk.Toplevel(parent)
        top.title("Seleccionar fecha")
        top.resizable(False, False)
        top.grab_set()
        top.configure(bg="white")

        if initial_date is None:
            initial_date = datetime.date.today()

        cal = Calendar(top, selectmode="day",
                       year=initial_date.year, month=initial_date.month, day=initial_date.day,
                       date_pattern="dd/mm/yyyy",
                       font=(FF, 11),
                       background=C["pri"], foreground="white",
                       headersbackground=C["pri"], headersforeground="white",
                       selectbackground=C["pri"], selectforeground="white",
                       normalbackground="white", normalforeground=C["tx"],
                       weekendbackground="#F5F6F8", weekendforeground=C["tx"],
                       othermonthbackground="#F5F6F8", othermonthforeground=C["tx3"],
                       othermonthwebackground="#F5F6F8", othermonthweforeground=C["tx3"],
                       borderwidth=0)
        cal.pack(padx=12, pady=(12,6))

        def confirm():
            self.result = cal.selection_get()
            self.entry_var.set(self.result.strftime("%d/%m/%Y"))
            top.destroy()

        btn_frame = tk.Frame(top, bg="white")
        btn_frame.pack(pady=(4,12))
        tk.Button(btn_frame, text="  Seleccionar  ", command=confirm,
                  bg=C["pri"], fg="white", font=(FF, 11, "bold"),
                  relief="flat", cursor="hand2", padx=16, pady=6).pack()

        # Center on parent
        top.update_idletasks()
        pw = parent.winfo_rootx() + parent.winfo_width()//2
        ph = parent.winfo_rooty() + parent.winfo_height()//2
        tw = top.winfo_width(); th = top.winfo_height()
        top.geometry(f"+{pw-tw//2}+{ph-th//2}")
        parent.wait_window(top)

# ═══════════════ POO ═══════════════
class Usuario:
    def __init__(s,n,c,e=""):
        s.id=str(uuid.uuid4())[:8]; s.nombre=n.strip().title()
        s.cargo=c.strip().title(); s.email=e.strip().lower()
        s.avatar_path=""; s.fecha_registro=datetime.datetime.now()
    def to_dict(s):
        return {"id":s.id,"nombre":s.nombre,"cargo":s.cargo,"email":s.email,
                "avatar_path":s.avatar_path,"fecha_registro":s.fecha_registro.isoformat()}
    @classmethod
    def from_dict(c,d):
        u=c(d["nombre"],d["cargo"],d.get("email",""))
        u.id=d["id"]; u.avatar_path=d.get("avatar_path","")
        u.fecha_registro=datetime.datetime.fromisoformat(d["fecha_registro"]); return u

class Tarea:
    ESTADOS=["Pendiente","En Progreso","Completada"]
    PRIORIDADES=["Baja","Media","Alta","Crítica"]
    def __init__(s,tit,desc,fl,rid,prio="Media"):
        s.id=str(uuid.uuid4())[:8]; s.titulo=tit.strip().title()
        s.descripcion=desc.strip(); s.fecha_creacion=datetime.datetime.now()
        s.fecha_limite=fl; s.estado="Pendiente"; s.prioridad=prio
        s.responsable_id=rid; s.historial=[]; s._log("Tarea creada")
    def _log(s,m): s.historial.append({"fecha":datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"cambio":m})
    def cambiar_estado(s,n):
        if n in s.ESTADOS: a=s.estado; s.estado=n; s._log(f"Estado: {a} → {n}")
    def modificar(s,**kw):
        for k,v in kw.items():
            if hasattr(s,k): a=getattr(s,k); setattr(s,k,v); s._log(f"{k}: {a} → {v}")
    def dias_restantes(s):
        h=datetime.date.today()
        fl=s.fecha_limite.date() if isinstance(s.fecha_limite,datetime.datetime) else s.fecha_limite
        return (fl-h).days
    def esta_vencida(s): return s.dias_restantes()<0 and s.estado!="Completada"
    def dias_total_plazo(s):
        fl=s.fecha_limite.date() if isinstance(s.fecha_limite,datetime.datetime) else s.fecha_limite
        return (fl-s.fecha_creacion.date()).days
    def porcentaje_sla(s):
        t=s.dias_total_plazo()
        if t<=0: return 100
        return min(100,max(0,int((t-s.dias_restantes())/t*100)))
    def siguiente_estado(s):
        if s.estado=="Pendiente": return "En Progreso"
        elif s.estado=="En Progreso": return "Completada"
        return None
    def to_dict(s):
        fl=s.fecha_limite
        if isinstance(fl,datetime.datetime): fl=fl.date()
        return {"id":s.id,"titulo":s.titulo,"descripcion":s.descripcion,
                "fecha_creacion":s.fecha_creacion.isoformat(),"fecha_limite":fl.isoformat(),
                "estado":s.estado,"prioridad":s.prioridad,"responsable_id":s.responsable_id,
                "historial":s.historial}
    @classmethod
    def from_dict(c,d):
        fl=datetime.date.fromisoformat(d["fecha_limite"])
        t=c(d["titulo"],d["descripcion"],fl,d["responsable_id"],d.get("prioridad","Media"))
        t.id=d["id"]; t.fecha_creacion=datetime.datetime.fromisoformat(d["fecha_creacion"])
        t.estado=d["estado"]; t.historial=d.get("historial",[]); return t

# ═══════════════ GESTOR ═══════════════
class GestorTareas:
    def __init__(s): s.usuarios=[]; s.tareas=[]; s.cargar_datos()
    def gen_por_estado(s,e):
        for t in s.tareas:
            if t.estado==e: yield t
    def gen_por_usuario(s,uid):
        for t in s.tareas:
            if t.responsable_id==uid: yield t
    def gen_vencidas(s):
        for t in s.tareas:
            if t.esta_vencida(): yield t
    def add_usuario(s,n,c,e="",av=""):
        u=Usuario(n,c,e); u.avatar_path=av; s.usuarios.append(u); s.guardar(); return u
    def del_usuario(s,uid):
        if any(t.estado!="Completada" for t in s.tareas if t.responsable_id==uid): return False
        s.usuarios=[u for u in s.usuarios if u.id!=uid]; s.guardar(); return True
    def add_tarea(s,ti,de,fl,rid,pr="Media"):
        t=Tarea(ti,de,fl,rid,pr); s.tareas.append(t)
        s.tareas.sort(key=lambda x:x.fecha_limite); s.guardar(); return t
    def del_tarea(s,tid): s.tareas=[t for t in s.tareas if t.id!=tid]; s.guardar()
    def get_usr(s,uid): return next((u for u in s.usuarios if u.id==uid),None)
    def get_nombre(s,uid): u=s.get_usr(uid); return u.nombre if u else "Sin asignar"
    def informe_txt(s):
        a=datetime.datetime.now(); l=[]
        l.append("="*70); l.append("  INFORME ORKESTIA"); l.append(f"  {a.strftime('%d/%m/%Y %H:%M:%S')}"); l.append("="*70)
        tot=len(s.tareas); pe=len([t for t in s.tareas if t.estado=="Pendiente"])
        pr=len([t for t in s.tareas if t.estado=="En Progreso"])
        co=len([t for t in s.tareas if t.estado=="Completada"])
        ve=len(list(s.gen_vencidas()))
        l.append("\n📊 RESUMEN")
        l.append(tabulate([["Total",tot],["Pendientes",pe],["En progreso",pr],["Completadas",co],
                           ["Vencidas",ve],["Usuarios",len(s.usuarios)]],
                          headers=["Concepto","Cant."],tablefmt="rounded_grid"))
        l.append("\n👥 POR USUARIO")
        for u in s.usuarios:
            tu=list(s.gen_por_usuario(u.id))
            if tu:
                l.append(f"\n  ► {u.nombre}")
                tb=[[t.titulo[:30],t.estado,t.prioridad,f"{t.dias_restantes()}d"] for t in tu]
                l.append(tabulate(tb,headers=["Tarea","Estado","Prio","Días"],tablefmt="rounded_grid"))
        l.append("\n📅 CALENDARIO"); cl=calendar.TextCalendar(calendar.MONDAY)
        l.append(cl.formatmonth(a.year,a.month)); l.append("="*70); return "\n".join(l)
    def exp_txt(s,r=None):
        if not r: r=RPTS/f"informe_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(r,"w",encoding="utf-8") as f: f.write(s.informe_txt()); return str(r)
    def exp_xlsx(s,r=None):
        if not r: r=RPTS/f"informe_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb=XlWb(); ws=wb.active; ws.title="Tareas"
        hd=["ID","Título","Responsable","Prioridad","Estado","F.Creación","F.Límite","Días","SLA%"]
        hf=XlF(bold=True,color="FFFFFF",name="Arial",size=11); hp=XlP("solid",fgColor="6C60FF")
        for i,h in enumerate(hd,1):
            c2=ws.cell(row=1,column=i,value=h); c2.font=hf; c2.fill=hp; c2.alignment=XlA(horizontal="center")
        for r2,t in enumerate(s.tareas,2):
            fl=t.fecha_limite
            if isinstance(fl,datetime.datetime): fl=fl.date()
            ws.cell(row=r2,column=1,value=t.id); ws.cell(row=r2,column=2,value=t.titulo)
            ws.cell(row=r2,column=3,value=s.get_nombre(t.responsable_id))
            ws.cell(row=r2,column=4,value=t.prioridad); ws.cell(row=r2,column=5,value=t.estado)
            ws.cell(row=r2,column=6,value=t.fecha_creacion.strftime("%d/%m/%Y"))
            ws.cell(row=r2,column=7,value=fl.strftime("%d/%m/%Y"))
            ws.cell(row=r2,column=8,value=t.dias_restantes()); ws.cell(row=r2,column=9,value=f"{t.porcentaje_sla()}%")
        for i in range(1,len(hd)+1): ws.column_dimensions[chr(64+i)].width=18
        wb.save(str(r)); return str(r)
    def guardar(s):
        d={"usuarios":[u.to_dict() for u in s.usuarios],"tareas":[t.to_dict() for t in s.tareas]}
        with open(TFILE,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
        with open(BFILE,"wb") as f: pickle.dump({"usuarios":s.usuarios,"tareas":s.tareas},f)
    def cargar_datos(s):
        if os.path.exists(BFILE):
            try:
                with open(BFILE,"rb") as f: d=pickle.load(f)
                s.usuarios=d.get("usuarios",[]); s.tareas=d.get("tareas",[]); return
            except: pass
        if os.path.exists(TFILE):
            try:
                with open(TFILE,"r",encoding="utf-8") as f: d=json.load(f)
                s.usuarios=[Usuario.from_dict(u) for u in d.get("usuarios",[])]
                s.tareas=[Tarea.from_dict(t) for t in d.get("tareas",[])]
            except: pass
    def backup(s):
        r=BKUP/f"backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pkl"
        with open(r,"wb") as f: pickle.dump({"usuarios":s.usuarios,"tareas":s.tareas},f)
        return str(r)

# ─── Helpers ───
def avatar_circular(path, size=40):
    try:
        img=Image.open(path).convert("RGBA").resize((size,size),Image.LANCZOS)
        mask=Image.new("L",(size,size),0); ImageDraw.Draw(mask).ellipse((0,0,size,size),fill=255)
        img.putalpha(mask); return ImageTk.PhotoImage(img)
    except: return None

def cargar_logo(path, w=120, h=120):
    try:
        img=Image.open(path).convert("RGBA"); img.thumbnail((w,h),Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except: return None

# ═══════════════ APP ═══════════════
class App(ctk.CTk):
    SB_W = 260; SB_C = 68

    def __init__(self):
        super().__init__()
        self.G = GestorTareas(); self._avc = {}
        self.title("Orkestia — Gestión de Tareas")
        self.geometry("1400x840"); self.minsize(1100,700)
        ctk.set_appearance_mode("light"); ctk.set_default_color_theme("blue")
        self.configure(fg_color=C["bg2"])
        self._sb_open = True
        self._build(); self._show_dash()

    def _f(s,sz=13,w="normal"): return ctk.CTkFont(family=FF,size=sz,weight=w)
    def _tt(s,w,t): Tooltip(w,t)

    def _avatar(s,usr,sz=36):
        if usr.avatar_path and os.path.exists(usr.avatar_path):
            k=f"{usr.id}_{sz}"
            if k not in s._avc: s._avc[k]=avatar_circular(usr.avatar_path,sz)
            return s._avc[k]
        return None

    def _ini_bubble(s, parent, usr, sz=28):
        """Draws an initials bubble for a user."""
        ab=ctk.CTkFrame(parent, width=sz, height=sz, fg_color=C["pril"], corner_radius=sz//2)
        ab.pack_propagate(False)
        ini="".join([p[0] for p in usr.nombre.split()[:2]]).upper()
        ctk.CTkLabel(ab, text=ini, font=s._f(max(8,sz//3),"bold"), text_color=C["pri"]).place(relx=0.5,rely=0.5,anchor="center")
        return ab

    # ── Build ──
    def _build(s):
        s._sbf = ctk.CTkFrame(s, fg_color=C["sb"], corner_radius=0,
                                border_width=1, border_color=C["brd"], width=s.SB_W)
        s._sbf.pack(side="left", fill="y"); s._sbf.pack_propagate(False)

        sb = ctk.CTkFrame(s._sbf, fg_color="transparent"); sb.pack(fill="both", expand=True)
        s._sb = sb

        # Logo — BIG (120px)
        logo_area = ctk.CTkFrame(sb, fg_color="transparent"); logo_area.pack(fill="x", padx=10, pady=(12,0))
        s._logo_img = None
        if LOGO.exists(): s._logo_img = cargar_logo(str(LOGO), 120, 120)
        if s._logo_img:
            s._logo_lbl = ctk.CTkLabel(logo_area, image=s._logo_img, text="", fg_color="transparent")
            s._logo_lbl.pack(anchor="center", pady=(0,2))

        s._name_lbl = ctk.CTkLabel(sb, text="Orkestia", font=s._f(22,"bold"), text_color=C["pri"])
        s._name_lbl.pack()
        s._slogan = ctk.CTkLabel(sb, text="Orquesta tu operación", font=s._f(9), text_color=C["tx3"])
        s._slogan.pack(pady=(0,4))

        s._tog = ctk.CTkButton(sb, text="◀  Contraer menú", fg_color=C["pril"],
                                text_color=C["pri"], hover_color=C["brd"],
                                font=s._f(11), height=32, corner_radius=8,
                                anchor="w", command=s._toggle)
        s._tog.pack(fill="x", padx=14, pady=(6,4)); s._tt(s._tog, "Contraer o expandir el menú lateral")

        ctk.CTkFrame(sb, height=1, fg_color=C["brd"]).pack(fill="x", padx=14, pady=(6,4))
        s._mlbl = ctk.CTkLabel(sb, text="  MENÚ", font=s._f(9,"bold"), text_color=C["tx3"])
        s._mlbl.pack(anchor="w", padx=18, pady=(2,4))

        menus=[("📊","Dashboard",s._show_dash,"Ver resumen general"),
               ("📋","Tareas",s._show_tasks,"Gestionar tareas"),
               ("👥","Usuarios",s._show_users,"Administrar usuarios"),
               ("📈","Informes",s._show_reports,"Reportes y exportar")]
        s._mbtn = []
        for ic,txt,cmd,tip in menus:
            b=ctk.CTkButton(sb, text=f"{ic}  {txt}", command=cmd,
                            fg_color="transparent", hover_color=C["pril"],
                            text_color=C["tx"], anchor="w",
                            font=s._f(13), height=40, corner_radius=10)
            b.pack(fill="x", padx=12, pady=2); s._mbtn.append(b); s._tt(b,tip)

        ctk.CTkFrame(sb, fg_color="transparent").pack(fill="both", expand=True)
        inf=ctk.CTkFrame(sb, fg_color=C["pris"], corner_radius=12); inf.pack(fill="x", padx=12, pady=(0,14))
        s._inf1=ctk.CTkLabel(inf, text="Desarrollado por", font=s._f(9,"bold"), text_color=C["pri"])
        s._inf1.pack(anchor="w", padx=10, pady=(8,0))
        s._inf2=ctk.CTkLabel(inf, text="Daniel Suarez", font=s._f(10,"bold"), text_color=C["tx"])
        s._inf2.pack(anchor="w", padx=10)
        s._inf3=ctk.CTkLabel(inf, text="Python para IA", font=s._f(9), text_color=C["tx2"])
        s._inf3.pack(anchor="w", padx=10)
        s._inf4=ctk.CTkLabel(inf, text="Maestría en Inteligencia Artificial", font=s._f(8), text_color=C["tx2"])
        s._inf4.pack(anchor="w", padx=10)
        s._inf5=ctk.CTkLabel(inf, text="Universidad de La Salle", font=s._f(8), text_color=C["tx3"])
        s._inf5.pack(anchor="w", padx=10, pady=(0,8))

        s.content = ctk.CTkFrame(s, fg_color=C["bg2"], corner_radius=0)
        s.content.pack(side="right", fill="both", expand=True)

    def _toggle(s):
        if s._sb_open:
            s._sb_target = s.SB_C; s._tog.configure(text="▶")
        else:
            s._sb_target = s.SB_W; s._tog.configure(text="◀  Contraer menú")
        s._sb_open = not s._sb_open
        s._anim()

    def _anim(s):
        cur = s._sbf.cget("width"); diff = s._sb_target - cur
        if abs(diff) < 3:
            s._sbf.configure(width=s._sb_target); s._upd_sb(); return
        step = int(diff*0.3) or (1 if diff>0 else -1)
        s._sbf.configure(width=cur+step); s.after(12, s._anim)

    def _upd_sb(s):
        if not s._sb_open:
            for w in [s._name_lbl, s._slogan, s._mlbl, s._inf1, s._inf2, s._inf3, s._inf4, s._inf5]:
                w.pack_forget()
            if hasattr(s,'_logo_lbl'): s._logo_lbl.pack_forget()
            labels=["📊","📋","👥","📈"]
            for b,l in zip(s._mbtn,labels): b.configure(text=l, anchor="center")
        else:
            if hasattr(s,'_logo_lbl') and s._logo_img:
                s._logo_lbl.pack(anchor="center", pady=(0,2))
            s._name_lbl.pack(); s._slogan.pack(pady=(0,4))
            s._mlbl.pack(anchor="w", padx=18, pady=(2,4))
            s._inf1.pack(anchor="w", padx=10, pady=(8,0))
            s._inf2.pack(anchor="w", padx=10)
            s._inf3.pack(anchor="w", padx=10)
            s._inf4.pack(anchor="w", padx=10)
            s._inf5.pack(anchor="w", padx=10, pady=(0,8))
            labels=["📊  Dashboard","📋  Tareas","👥  Usuarios","📈  Informes"]
            for b,l in zip(s._mbtn,labels): b.configure(text=l, anchor="w")

    def _cl(s):
        for w in s.content.winfo_children(): w.destroy()
    def _hl(s,i):
        for j,b in enumerate(s._mbtn):
            b.configure(fg_color=C["pril"] if j==i else "transparent",
                        text_color=C["pri"] if j==i else C["tx"])

    # ═══════════ DASHBOARD ═══════════
    def _show_dash(s):
        s._cl(); s._hl(0)
        sc=ctk.CTkScrollableFrame(s.content, fg_color="transparent", scrollbar_button_color=C["brd"])
        sc.pack(fill="both", expand=True)
        hf=ctk.CTkFrame(sc, fg_color="transparent"); hf.pack(fill="x", padx=28, pady=(22,4))
        ctk.CTkLabel(hf, text="Dashboard", font=s._f(26,"bold"), text_color=C["tx"]).pack(anchor="w")
        ctk.CTkLabel(hf, text="Resumen general de operación", font=s._f(12), text_color=C["tx3"]).pack(anchor="w")

        cf=ctk.CTkFrame(sc, fg_color="transparent"); cf.pack(fill="x", padx=28, pady=(12,8))
        tot=len(s.G.tareas); pe=len([t for t in s.G.tareas if t.estado=="Pendiente"])
        pr=len([t for t in s.G.tareas if t.estado=="En Progreso"])
        co=len([t for t in s.G.tareas if t.estado=="Completada"])
        ve=len(list(s.G.gen_vencidas()))
        for i,(lb,vl,co2,bg) in enumerate([
            ("Total",str(tot),C["pri"],C["pril"]),("Pendientes",str(pe),C["sp"],C["wal"]),
            ("En Progreso",str(pr),C["sg"],C["infl"]),("Completadas",str(co),C["sc"],C["okl"]),
            ("Vencidas",str(ve),C["er"],C["erl"]),("Usuarios",str(len(s.G.usuarios)),C["pri"],C["pris"])]):
            cf.columnconfigure(i, weight=1)
            cd=ctk.CTkFrame(cf, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"], height=90)
            cd.grid(row=0,column=i,padx=3,pady=3,sticky="nsew"); cd.grid_propagate(False)
            ctk.CTkLabel(cd, text=vl, font=s._f(24,"bold"), text_color=co2).place(relx=0.5,rely=0.38,anchor="center")
            ctk.CTkLabel(cd, text=lb, font=s._f(10), text_color=C["tx3"]).place(relx=0.5,rely=0.72,anchor="center")

        ch=ctk.CTkFrame(sc, fg_color="transparent"); ch.pack(fill="x", padx=28, pady=(8,8))
        ch.columnconfigure(0, weight=3); ch.columnconfigure(1, weight=2)

        bc=ctk.CTkFrame(ch, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"])
        bc.grid(row=0,column=0,padx=(0,5),pady=3,sticky="nsew")
        ctk.CTkLabel(bc, text="Tareas por usuario", font=s._f(14,"bold"), text_color=C["tx"]).pack(anchor="w", padx=16, pady=(12,2))
        bcv=Canvas(bc, bg="#FFFFFF", highlightthickness=0, height=180); bcv.pack(fill="x", padx=14, pady=(2,12))
        s._draw_bars(bcv)

        pc=ctk.CTkFrame(ch, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"])
        pc.grid(row=0,column=1,padx=(5,0),pady=3,sticky="nsew")
        ctk.CTkLabel(pc, text="Estado de tickets", font=s._f(14,"bold"), text_color=C["tx"]).pack(anchor="w", padx=16, pady=(12,2))
        pcv=Canvas(pc, bg="#FFFFFF", highlightthickness=0, height=180); pcv.pack(fill="x", padx=14, pady=(2,12))
        s._draw_pie(pcv, pe, pr, co, ve)

        uc=ctk.CTkFrame(sc, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"])
        uc.pack(fill="x", padx=28, pady=(4,18))
        ctk.CTkLabel(uc, text="Usuarios y carga", font=s._f(14,"bold"), text_color=C["tx"]).pack(anchor="w", padx=16, pady=(12,6))
        if not s.G.usuarios:
            ctk.CTkLabel(uc, text="Sin usuarios", font=s._f(12), text_color=C["tx3"]).pack(pady=16)
        else:
            for u in s.G.usuarios:
                r=ctk.CTkFrame(uc, fg_color="transparent"); r.pack(fill="x", padx=16, pady=2)
                av=s._avatar(u,28)
                if av: ctk.CTkLabel(r, image=av, text="", fg_color="transparent").pack(side="left", padx=(0,8))
                else: s._ini_bubble(r, u, 28).pack(side="left", padx=(0,8))
                ctk.CTkLabel(r, text=u.nombre, font=s._f(12,"bold"), text_color=C["tx"]).pack(side="left")
                nt=len(list(s.G.gen_por_usuario(u.id)))
                ctk.CTkLabel(r, text=f"  {nt} tareas", font=s._f(10), text_color=C["tx3"]).pack(side="left", padx=6)

    def _draw_bars(s, cv):
        cv.update_idletasks(); w=max(cv.winfo_width(),380); h=170; cv.delete("all")
        if not s.G.usuarios: cv.create_text(w//2,h//2,text="Sin datos",fill=C["tx3"],font=(FF,11)); return
        datos=[]
        for u in s.G.usuarios:
            tu=list(s.G.gen_por_usuario(u.id))
            datos.append((u.nombre.split()[0],
                          len([t for t in tu if t.estado=="Pendiente"]),
                          len([t for t in tu if t.estado=="En Progreso"]),
                          len([t for t in tu if t.estado=="Completada"])))
        n=len(datos); bw=min(45,(w-80)//(n*1.5)); gap=bw//2
        mx=max(sum(d[1:]) for d in datos) or 1; left=50
        for i,(nm,pe,pr,co) in enumerate(datos):
            x=left+i*(bw+gap); yb=h-18
            for val,color in [(co,C["sc"]),(pr,C["sg"]),(pe,C["sp"])]:
                bh=int(val/mx*(h-45)) if mx>0 else 0
                if bh>0: cv.create_rectangle(x,yb-bh,x+bw,yb,fill=color,outline=""); yb-=bh
            cv.create_text(x+bw//2,h-5,text=nm[:7],fill=C["tx2"],font=(FF,8))
        lx=w-180
        for j,(lb,co) in enumerate([("Pendiente",C["sp"]),("En Progreso",C["sg"]),("Completada",C["sc"])]):
            cv.create_rectangle(lx,8+j*16,lx+10,18+j*16,fill=co,outline="")
            cv.create_text(lx+16,13+j*16,text=lb,anchor="w",fill=C["tx2"],font=(FF,8))

    def _draw_pie(s, cv, pe, pr, co, ve):
        cv.update_idletasks(); w=max(cv.winfo_width(),230); h=170; cv.delete("all")
        tot=pe+pr+co+ve
        if tot==0: cv.create_text(w//2,h//2,text="Sin datos",fill=C["tx3"],font=(FF,11)); return
        cx,cy,r=w//2-35,h//2,62
        slices=[("Pend.",pe,C["sp"]),("Progr.",pr,C["sg"]),("Compl.",co,C["sc"]),("Venc.",ve,C["er"])]
        st=0
        for lb,val,co2 in slices:
            if val==0: continue
            ext=val/tot*360
            cv.create_arc(cx-r,cy-r,cx+r,cy+r,start=st,extent=ext,fill=co2,outline=C["cd"],width=2)
            mid=math.radians(st+ext/2)
            tx2=cx+(r+18)*math.cos(mid); ty=cy-(r+18)*math.sin(mid)
            cv.create_text(tx2,ty,text=f"{int(val/tot*100)}%",fill=co2,font=(FF,8,"bold")); st+=ext
        lx=cx+r+45
        for j,(lb,val,co2) in enumerate(slices):
            if val==0: continue
            cv.create_rectangle(lx,25+j*20,lx+12,37+j*20,fill=co2,outline="")
            cv.create_text(lx+18,31+j*20,text=f"{lb} ({val})",anchor="w",fill=C["tx2"],font=(FF,9))

    # ═══════════ TAREAS ═══════════
    def _show_tasks(s):
        s._cl(); s._hl(1)
        hf=ctk.CTkFrame(s.content, fg_color="transparent"); hf.pack(fill="x", padx=28, pady=(22,4))
        ctk.CTkLabel(hf, text="Gestión de Tareas", font=s._f(26,"bold"), text_color=C["tx"]).pack(side="left")
        nb=ctk.CTkButton(hf, text="＋ Nueva Tarea", fg_color=C["pri"], text_color=C["wh"],
                         hover_color=C["prih"], font=s._f(12,"bold"), corner_radius=10,
                         height=36, width=150, command=s._dlg_new_task)
        nb.pack(side="right"); s._tt(nb, "Crear nueva tarea")

        # ── Filters — uniform row ──
        fi=ctk.CTkFrame(s.content, fg_color=C["cd"], corner_radius=12,
                         border_width=1, border_color=C["brdl"])
        fi.pack(fill="x", padx=28, pady=(10,6))
        fi_row=ctk.CTkFrame(fi, fg_color="transparent"); fi_row.pack(fill="x", padx=14, pady=10)

        # Estado
        ef=ctk.CTkFrame(fi_row, fg_color="transparent"); ef.pack(side="left", padx=(0,12))
        ctk.CTkLabel(ef, text="Estado", font=s._f(9,"bold"), text_color=C["tx3"]).pack(anchor="w")
        s._fe=ctk.CTkComboBox(ef, values=["Todos","Pendiente","En Progreso","Completada"],
                               width=140, height=32, font=s._f(11), fg_color=C["inp"],
                               border_color=C["brd"], button_color=C["pri"],
                               button_hover_color=C["prih"],
                               command=lambda _: s._filter())
        s._fe.set("Todos"); s._fe.pack(pady=(2,0)); s._tt(s._fe, "Filtrar por estado")

        # Usuario
        uf=ctk.CTkFrame(fi_row, fg_color="transparent"); uf.pack(side="left", padx=(0,12))
        ctk.CTkLabel(uf, text="Usuario", font=s._f(9,"bold"), text_color=C["tx3"]).pack(anchor="w")
        un=["Todos"]+[u.nombre for u in s.G.usuarios]
        s._fu=ctk.CTkComboBox(uf, values=un, width=155, height=32, font=s._f(11),
                               fg_color=C["inp"], border_color=C["brd"], button_color=C["pri"],
                               button_hover_color=C["prih"],
                               command=lambda _: s._filter())
        s._fu.set("Todos"); s._fu.pack(pady=(2,0)); s._tt(s._fu, "Filtrar por responsable")

        # Vencimiento — uniform with others
        vf=ctk.CTkFrame(fi_row, fg_color="transparent"); vf.pack(side="left", padx=(0,12))
        ctk.CTkLabel(vf, text="Vencimiento hasta", font=s._f(9,"bold"), text_color=C["tx3"]).pack(anchor="w")

        vf_row = ctk.CTkFrame(vf, fg_color="transparent"); vf_row.pack(pady=(2,0))
        s._fdate_var = tk.StringVar(value="")
        s._fdate_entry = ctk.CTkEntry(vf_row, textvariable=s._fdate_var, width=105, height=32,
                                       font=s._f(11), fg_color=C["inp"], border_color=C["brd"],
                                       placeholder_text="dd/mm/aaaa")
        s._fdate_entry.pack(side="left")

        dcb = ctk.CTkButton(vf_row, text="📅", width=32, height=32, fg_color=C["pri"],
                            text_color="white", hover_color=C["prih"], corner_radius=8,
                            font=s._f(12), command=lambda: s._pick_filter_date())
        dcb.pack(side="left", padx=(4,0)); s._tt(dcb, "Abrir calendario para seleccionar fecha")

        s._fdate_active = tk.BooleanVar(value=False)
        cb=ctk.CTkCheckBox(vf, text="Activo", font=s._f(9), text_color=C["tx3"],
                            variable=s._fdate_active, width=20, height=20,
                            checkbox_width=16, checkbox_height=16,
                            fg_color=C["pri"], hover_color=C["prih"],
                            command=s._filter)
        cb.pack(anchor="w", pady=(3,0)); s._tt(cb, "Activar filtro por fecha de vencimiento")

        # Buscar
        sf=ctk.CTkFrame(fi_row, fg_color="transparent"); sf.pack(side="right", padx=(12,0))
        ctk.CTkLabel(sf, text="Buscar", font=s._f(9,"bold"), text_color=C["tx3"]).pack(anchor="w")
        s._fs=ctk.StringVar(); s._fs.trace_add("write", lambda *_: s._filter())
        se=ctk.CTkEntry(sf, textvariable=s._fs, placeholder_text="🔍 Buscar...",
                        width=180, height=32, font=s._f(11), fg_color=C["inp"], border_color=C["brd"])
        se.pack(pady=(2,0)); s._tt(se, "Buscar por título o descripción")

        # Table header
        th=ctk.CTkFrame(s.content, fg_color=C["pri"], corner_radius=8, height=34)
        th.pack(fill="x", padx=28, pady=(4,0)); th.pack_propagate(False)
        for nm,w2 in [("",30),("Actividad",180),("Responsable",115),("Prioridad",70),
                      ("Creación",75),("Vencimiento",80),("Estado",85),("SLA / Plazo",150)]:
            ctk.CTkLabel(th, text=nm, width=w2, font=s._f(10,"bold"), text_color=C["wh"]).pack(side="left", padx=2)

        s._tscr=ctk.CTkScrollableFrame(s.content, fg_color="transparent", scrollbar_button_color=C["brd"])
        s._tscr.pack(fill="both", expand=True, padx=28, pady=(0,18))
        s._filter()

    def _pick_filter_date(s):
        try:
            parts = s._fdate_var.get().strip().split("/")
            d = datetime.date(int(parts[2]),int(parts[1]),int(parts[0]))
        except: d = datetime.date.today()
        CalendarPopup(s, s._fdate_var, d)

    def _filter(s):
        for w in s._tscr.winfo_children(): w.destroy()
        e=s._fe.get(); u=s._fu.get(); b=s._fs.get().lower().strip()
        tareas=s.G.tareas.copy()
        if e!="Todos": tareas=[t for t in tareas if t.estado==e]
        if u!="Todos":
            uid=next((x.id for x in s.G.usuarios if x.nombre==u),None)
            if uid: tareas=[t for t in tareas if t.responsable_id==uid]
        if b: tareas=[t for t in tareas if b in t.titulo.lower() or b in t.descripcion.lower()]
        if s._fdate_active.get():
            try:
                parts = s._fdate_var.get().strip().split("/")
                fd = datetime.date(int(parts[2]),int(parts[1]),int(parts[0]))
                tareas=[t for t in tareas if (t.fecha_limite.date() if isinstance(t.fecha_limite,datetime.datetime) else t.fecha_limite) <= fd]
            except: pass
        if not tareas:
            ctk.CTkLabel(s._tscr, text="No se encontraron tareas", font=s._f(13), text_color=C["tx3"]).pack(pady=35)
            return
        for t in tareas: s._task_row(s._tscr, t)

    def _task_row(s, parent, t):
        ce={"Pendiente":C["sp"],"En Progreso":C["sg"],"Completada":C["sc"]}.get(t.estado,C["tx3"])
        cp={"Crítica":C["pc"],"Alta":C["pa"],"Media":C["pm"],"Baja":C["pb"]}.get(t.prioridad,C["tx3"])
        row=ctk.CTkFrame(parent, fg_color=C["cd"], corner_radius=8,
                          border_width=1, border_color=C["brdl"], height=40)
        row.pack(fill="x", pady=2); row.pack_propagate(False)

        ctk.CTkFrame(row, width=4, fg_color=ce, corner_radius=2).pack(side="left", fill="y", padx=(5,4), pady=5)

        usr=s.G.get_usr(t.responsable_id)
        if usr:
            av=s._avatar(usr,24)
            if av: ctk.CTkLabel(row, image=av, text="", fg_color="transparent").pack(side="left", padx=(0,4))
            else: s._ini_bubble(row, usr, 24).pack(side="left", padx=(0,4))

        ctk.CTkLabel(row, text=t.titulo[:25], font=s._f(11,"bold"), text_color=C["tx"], anchor="w", width=180).pack(side="left", padx=3)
        ctk.CTkLabel(row, text=s.G.get_nombre(t.responsable_id)[:14], font=s._f(10), text_color=C["tx2"], anchor="w", width=115).pack(side="left", padx=3)
        ctk.CTkLabel(row, text=f" {t.prioridad} ", font=s._f(9,"bold"), fg_color=cp, text_color="white", corner_radius=4, width=70).pack(side="left", padx=3)
        ctk.CTkLabel(row, text=t.fecha_creacion.strftime("%d/%m/%y"), font=s._f(9), text_color=C["tx3"], width=75).pack(side="left", padx=3)
        fl=t.fecha_limite
        if isinstance(fl,datetime.datetime): fl=fl.date()
        ctk.CTkLabel(row, text=fl.strftime("%d/%m/%y"), font=s._f(9), text_color=C["tx2"], width=80).pack(side="left", padx=3)
        ctk.CTkLabel(row, text=f" {t.estado} ", font=s._f(9,"bold"), fg_color=ce, text_color="white", corner_radius=4, width=85).pack(side="left", padx=3)

        # SLA heatmap
        dias=t.dias_restantes()
        if t.estado=="Completada": sla_c=C["sc"]; dtx="✅"
        elif dias<0: sla_c="#B80020"; dtx=f"{dias}d"
        elif dias<=3: sla_c=C["er"]; dtx=f"{dias}d"
        elif dias<=6: sla_c=C["wa"]; dtx=f"{dias}d"
        else: sla_c=C["ok"]; dtx=f"{dias}d"

        sf=ctk.CTkFrame(row, fg_color=C["inp"], corner_radius=5, height=22, width=120)
        sf.pack(side="left", padx=6, pady=9); sf.pack_propagate(False)
        pct=min(100,t.porcentaje_sla()) if t.estado!="Completada" else 100
        bw2=max(22,int(120*pct/100))
        bar=ctk.CTkFrame(sf, width=bw2, fg_color=sla_c, corner_radius=5)
        bar.pack(side="left", fill="y"); bar.pack_propagate(False)
        ctk.CTkLabel(bar, text=dtx, font=s._f(9,"bold"), text_color="white").place(relx=0.5,rely=0.5,anchor="center")

        bf=ctk.CTkFrame(row, fg_color="transparent"); bf.pack(side="right", padx=5)
        if t.estado!="Completada":
            ne=t.siguiente_estado()
            sb2=ctk.CTkButton(bf, text="Siguiente Estado ▸", width=130, height=24,
                              fg_color=C["pri"], text_color="white", hover_color=C["prih"],
                              corner_radius=6, font=s._f(9,"bold"),
                              command=lambda ta=t,e=ne: s._chg(ta,e))
            sb2.pack(side="left", padx=1); s._tt(sb2, f"Cambiar a: {ne}")

        eb=ctk.CTkButton(bf, text="✏️", width=24, height=24, fg_color=C["inp"],
                         text_color=C["tx"], hover_color=C["brd"], corner_radius=6,
                         command=lambda ta=t: s._form_task(ta))
        eb.pack(side="left", padx=1); s._tt(eb, "Editar tarea")
        db=ctk.CTkButton(bf, text="🗑️", width=24, height=24, fg_color=C["erl"],
                         text_color=C["er"], hover_color=C["er"], corner_radius=6,
                         command=lambda ta=t: s._del_task(ta))
        db.pack(side="left", padx=1); s._tt(db, "Eliminar tarea")

    def _chg(s,t,e):
        if e: t.cambiar_estado(e); s.G.guardar(); s._show_tasks()
    def _del_task(s,t):
        if messagebox.askyesno("Confirmar",f"¿Eliminar '{t.titulo}'?"):
            s.G.del_tarea(t.id); s._show_tasks()

    def _dlg_new_task(s):
        if not s.G.usuarios:
            messagebox.showwarning("","Agrega al menos un usuario."); s._show_users(); return
        s._form_task()

    def _form_task(s, tarea=None):
        dlg=ctk.CTkToplevel(s); dlg.title("Editar Tarea" if tarea else "Nueva Tarea")
        dlg.geometry("500x600"); dlg.configure(fg_color=C["bg"])
        dlg.transient(s); dlg.grab_set(); dlg.after(100, dlg.lift)

        fm=ctk.CTkScrollableFrame(dlg, fg_color="transparent"); fm.pack(fill="both", expand=True, padx=26, pady=20)
        ctk.CTkLabel(fm, text="Editar Tarea" if tarea else "Nueva Tarea",
                     font=s._f(20,"bold"), text_color=C["pri"]).pack(anchor="w", pady=(0,14))

        ctk.CTkLabel(fm, text="Título *", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        te=ctk.CTkEntry(fm, height=36, font=s._f(11), fg_color=C["inp"],
                         border_color=C["brd"], placeholder_text="Nombre de la tarea")
        te.pack(fill="x", pady=(2,10)); s._tt(te, "Título de la tarea")
        if tarea: te.insert(0, tarea.titulo)

        ctk.CTkLabel(fm, text="Descripción", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        dt=ctk.CTkTextbox(fm, height=70, font=s._f(11), fg_color=C["inp"],
                           border_color=C["brd"], border_width=1)
        dt.pack(fill="x", pady=(2,10))
        if tarea: dt.insert("1.0", tarea.descripcion)

        # Fecha creación (readonly)
        ctk.CTkLabel(fm, text="Fecha de Creación", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        fc_txt=tarea.fecha_creacion.strftime("%d/%m/%Y %H:%M") if tarea else datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        fc_lbl=ctk.CTkLabel(fm, text=f"📅  {fc_txt}", font=s._f(11), text_color=C["tx"],
                             fg_color=C["inp"], corner_radius=8, height=34, anchor="w")
        fc_lbl.pack(fill="x", pady=(2,10)); s._tt(fc_lbl, "Fecha automática (no editable)")

        # Fecha Límite — Calendar Popup
        ctk.CTkLabel(fm, text="Fecha Límite *", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        fd=tarea.fecha_limite if tarea else datetime.date.today()+datetime.timedelta(days=7)
        if isinstance(fd, datetime.datetime): fd=fd.date()

        date_var = tk.StringVar(value=fd.strftime("%d/%m/%Y"))
        date_row = ctk.CTkFrame(fm, fg_color="transparent"); date_row.pack(fill="x", pady=(2,10))
        de = ctk.CTkEntry(date_row, textvariable=date_var, width=140, height=36,
                          font=s._f(11), fg_color=C["inp"], border_color=C["brd"])
        de.pack(side="left")

        def open_cal():
            try:
                parts=date_var.get().strip().split("/")
                d=datetime.date(int(parts[2]),int(parts[1]),int(parts[0]))
            except: d=datetime.date.today()
            CalendarPopup(dlg, date_var, d)

        cal_btn = ctk.CTkButton(date_row, text="📅 Elegir fecha", width=120, height=36,
                                fg_color=C["pri"], text_color="white", hover_color=C["prih"],
                                font=s._f(11), corner_radius=8, command=open_cal)
        cal_btn.pack(side="left", padx=(8,0)); s._tt(cal_btn, "Abrir calendario para seleccionar fecha")

        # Responsable
        ctk.CTkLabel(fm, text="Responsable *", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        um={u.nombre:u.id for u in s.G.usuarios}
        rc=ctk.CTkComboBox(fm, values=list(um.keys()), height=36, font=s._f(11),
                            fg_color=C["inp"], border_color=C["brd"], button_color=C["pri"],
                            button_hover_color=C["prih"])
        rc.pack(fill="x", pady=(2,10)); s._tt(rc, "Selecciona el responsable")
        if tarea: rc.set(s.G.get_nombre(tarea.responsable_id))

        ctk.CTkLabel(fm, text="Prioridad", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
        pc=ctk.CTkComboBox(fm, values=Tarea.PRIORIDADES, height=36, font=s._f(11),
                            fg_color=C["inp"], border_color=C["brd"], button_color=C["pri"])
        pc.pack(fill="x", pady=(2,10)); pc.set(tarea.prioridad if tarea else "Media")
        s._tt(pc, "Nivel de prioridad")

        ec=None
        if tarea:
            ctk.CTkLabel(fm, text="Estado", font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
            ec=ctk.CTkComboBox(fm, values=Tarea.ESTADOS, height=36, font=s._f(11),
                                fg_color=C["inp"], border_color=C["brd"], button_color=C["pri"])
            ec.set(tarea.estado); ec.pack(fill="x", pady=(2,10))

        def save():
            ti=te.get().strip(); ds=dt.get("1.0","end-1c").strip()
            rn=rc.get(); pr=pc.get()
            if not ti: messagebox.showwarning("","Título obligatorio."); return
            if rn not in um: messagebox.showwarning("","Responsable inválido."); return
            try:
                parts=date_var.get().strip().split("/")
                fe=datetime.date(int(parts[2]),int(parts[1]),int(parts[0]))
            except:
                messagebox.showwarning("","Fecha límite inválida. Use dd/mm/aaaa."); return
            if tarea:
                tarea.modificar(titulo=ti,descripcion=ds,fecha_limite=fe,responsable_id=um[rn],prioridad=pr)
                if ec: tarea.cambiar_estado(ec.get())
                s.G.guardar()
            else: s.G.add_tarea(ti,ds,fe,um[rn],pr)
            dlg.destroy(); s._show_tasks()

        sb2=ctk.CTkButton(fm, text="💾  Guardar", fg_color=C["pri"], text_color="white",
                          hover_color=C["prih"], font=s._f(12,"bold"), height=40,
                          corner_radius=10, command=save)
        sb2.pack(fill="x", pady=(14,0)); s._tt(sb2, "Guardar tarea")

    # ═══════════ USUARIOS ═══════════
    def _show_users(s):
        s._cl(); s._hl(2)
        hf=ctk.CTkFrame(s.content, fg_color="transparent"); hf.pack(fill="x", padx=28, pady=(22,4))
        ctk.CTkLabel(hf, text="Gestión de Usuarios", font=s._f(26,"bold"), text_color=C["tx"]).pack(side="left")
        nb=ctk.CTkButton(hf, text="＋ Nuevo Usuario", fg_color=C["pri"], text_color="white",
                         hover_color=C["prih"], font=s._f(12,"bold"), corner_radius=10,
                         height=36, width=160, command=s._dlg_new_user)
        nb.pack(side="right"); s._tt(nb, "Agregar nuevo usuario")
        sc=ctk.CTkScrollableFrame(s.content, fg_color="transparent", scrollbar_button_color=C["brd"])
        sc.pack(fill="both", expand=True, padx=28, pady=(10,18))
        if not s.G.usuarios:
            ctk.CTkLabel(sc, text="👥\nSin usuarios", font=s._f(14), text_color=C["tx3"], justify="center").pack(pady=40)
            return
        for u in s.G.usuarios:
            tu=list(s.G.gen_por_usuario(u.id))
            co=len([t for t in tu if t.estado=="Completada"]); ac=len([t for t in tu if t.estado!="Completada"])
            cd=ctk.CTkFrame(sc, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"])
            cd.pack(fill="x", pady=3)
            inn=ctk.CTkFrame(cd, fg_color="transparent"); inn.pack(fill="x", padx=16, pady=10)
            av=s._avatar(u,44)
            if av: ctk.CTkLabel(inn, image=av, text="", fg_color="transparent").pack(side="left", padx=(0,12))
            else: s._ini_bubble(inn, u, 44).pack(side="left", padx=(0,12))
            lf=ctk.CTkFrame(inn, fg_color="transparent"); lf.pack(side="left", fill="both", expand=True)
            ctk.CTkLabel(lf, text=u.nombre, font=s._f(14,"bold"), text_color=C["tx"]).pack(anchor="w")
            ctk.CTkLabel(lf, text=f"{u.cargo} · {u.email}", font=s._f(10), text_color=C["tx2"]).pack(anchor="w")
            sf=ctk.CTkFrame(lf, fg_color="transparent"); sf.pack(anchor="w", pady=(3,0))
            for lb,vl,co2 in [("Total",len(tu),C["pri"]),("Activas",ac,C["sg"]),("Listas",co,C["sc"])]:
                p=ctk.CTkFrame(sf, fg_color="transparent"); p.pack(side="left", padx=(0,12))
                ctk.CTkLabel(p, text=str(vl), font=s._f(12,"bold"), text_color=co2).pack(side="left", padx=(0,2))
                ctk.CTkLabel(p, text=lb, font=s._f(9), text_color=C["tx3"]).pack(side="left")
            rf=ctk.CTkFrame(inn, fg_color="transparent"); rf.pack(side="right")
            db=ctk.CTkButton(rf, text="🗑️ Eliminar", width=90, height=28, fg_color=C["erl"],
                             text_color=C["er"], hover_color=C["er"], border_width=1,
                             border_color=C["er"], font=s._f(9), corner_radius=8,
                             command=lambda us=u: s._del_user(us))
            db.pack(); s._tt(db, f"Eliminar a {u.nombre}")

    def _del_user(s,u):
        if messagebox.askyesno("Confirmar",f"¿Eliminar a '{u.nombre}'?"):
            if not s.G.del_usuario(u.id): messagebox.showwarning("","Tiene tareas activas.")
            else: s._show_users()

    def _dlg_new_user(s):
        dlg=ctk.CTkToplevel(s); dlg.title("Nuevo Usuario")
        dlg.geometry("450x470"); dlg.configure(fg_color=C["bg"]); dlg.transient(s); dlg.grab_set()
        dlg.after(100, dlg.lift)
        fm=ctk.CTkFrame(dlg, fg_color="transparent"); fm.pack(fill="both", expand=True, padx=26, pady=20)
        ctk.CTkLabel(fm, text="Nuevo Usuario", font=s._f(20,"bold"), text_color=C["pri"]).pack(anchor="w", pady=(0,14))
        avf=ctk.CTkFrame(fm, fg_color="transparent"); avf.pack(fill="x", pady=(0,10))
        s._nav=""
        s._avp=ctk.CTkLabel(avf, text="👤", font=s._f(28), fg_color=C["pril"], width=60, height=60, corner_radius=30)
        s._avp.pack(side="left", padx=(0,10))
        def sel_foto():
            p=filedialog.askopenfilename(filetypes=[("Imágenes","*.png *.jpg *.jpeg *.bmp")])
            if p:
                d2=AVDIR/f"{uuid.uuid4().hex[:8]}{Path(p).suffix}"
                shutil.copy2(p,d2); s._nav=str(d2)
                img=avatar_circular(str(d2),60)
                if img: s._avi=img; s._avp.configure(image=img, text="")
        sfb=ctk.CTkButton(avf, text="📷 Foto de perfil", fg_color=C["inp"], text_color=C["tx"],
                          hover_color=C["brd"], border_width=1, border_color=C["brd"],
                          font=s._f(10), height=32, corner_radius=8, command=sel_foto)
        sfb.pack(side="left"); s._tt(sfb, "Seleccionar foto de perfil")
        entries={}
        for lb,ph in [("Nombre completo *","Ej: María García"),("Cargo *","Ej: Desarrolladora"),("Email","Ej: maria@empresa.com")]:
            ctk.CTkLabel(fm, text=lb, font=s._f(10,"bold"), text_color=C["tx2"]).pack(anchor="w")
            e=ctk.CTkEntry(fm, height=36, placeholder_text=ph, font=s._f(11), fg_color=C["inp"], border_color=C["brd"])
            e.pack(fill="x", pady=(2,10)); entries[lb]=e
        def save():
            n=entries["Nombre completo *"].get().strip(); c=entries["Cargo *"].get().strip()
            em=entries["Email"].get().strip()
            if not n or not c: messagebox.showwarning("","Nombre y cargo obligatorios."); return
            s.G.add_usuario(n,c,em,s._nav); dlg.destroy(); s._show_users()
        sb2=ctk.CTkButton(fm, text="💾 Guardar", fg_color=C["pri"], text_color="white",
                          hover_color=C["prih"], font=s._f(12,"bold"), height=40, corner_radius=10, command=save)
        sb2.pack(fill="x", pady=(14,0)); s._tt(sb2, "Guardar usuario")

    # ═══════════ INFORMES ═══════════
    def _show_reports(s):
        s._cl(); s._hl(3)
        hf=ctk.CTkFrame(s.content, fg_color="transparent"); hf.pack(fill="x", padx=28, pady=(22,4))
        ctk.CTkLabel(hf, text="Informes y Reportes", font=s._f(26,"bold"), text_color=C["tx"]).pack(anchor="w")
        ctk.CTkLabel(hf, text="Genera, exporta y descarga reportes", font=s._f(12), text_color=C["tx3"]).pack(anchor="w")
        af=ctk.CTkFrame(s.content, fg_color="transparent"); af.pack(fill="x", padx=28, pady=(14,10))
        b1=ctk.CTkButton(af, text="📄 Exportar TXT", fg_color=C["pri"], text_color="white",
                         hover_color=C["prih"], font=s._f(11,"bold"), height=36, corner_radius=10, width=160,
                         command=s._etxt); b1.pack(side="left", padx=(0,6)); s._tt(b1, "Exportar como texto plano")
        b2=ctk.CTkButton(af, text="💾 Crear Backup", fg_color=C["sg"], text_color="white",
                         hover_color="#4A8AE0", font=s._f(11,"bold"), height=36, corner_radius=10, width=145,
                         command=s._ebkup); b2.pack(side="left", padx=(0,6)); s._tt(b2, "Copia de seguridad completa")
        b3=ctk.CTkButton(af, text="📊 Descargar Excel", fg_color=C["ok"], text_color="white",
                         hover_color="#00B060", font=s._f(11,"bold"), height=36, corner_radius=10, width=165,
                         command=s._exlsx); b3.pack(side="left"); s._tt(b3, "Descargar en formato .xlsx")
        pv=ctk.CTkFrame(s.content, fg_color=C["cd"], corner_radius=12, border_width=1, border_color=C["brdl"])
        pv.pack(fill="both", expand=True, padx=28, pady=(0,22))
        ctk.CTkLabel(pv, text="Vista previa", font=s._f(15,"bold"), text_color=C["tx"]).pack(anchor="w", padx=18, pady=(12,4))
        tb=ctk.CTkTextbox(pv, fg_color=C["inp"], font=ctk.CTkFont(family="Consolas",size=11),
                           text_color=C["tx"], border_width=1, border_color=C["brd"], corner_radius=10)
        tb.pack(fill="both", expand=True, padx=14, pady=(0,14))
        tb.insert("1.0", s.G.informe_txt()); tb.configure(state="disabled")

    def _etxt(s): r=s.G.exp_txt(); messagebox.showinfo("","Guardado en:\n"+r)
    def _ebkup(s): r=s.G.backup(); messagebox.showinfo("","Backup en:\n"+r)
    def _exlsx(s): r=s.G.exp_xlsx(); messagebox.showinfo("","Excel en:\n"+r)


if __name__=="__main__":
    app=App(); app.mainloop()
